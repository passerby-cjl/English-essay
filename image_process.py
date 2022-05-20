#!/usr/bin/python
#-*- coding:utf-8 -*-
import numpy as np
import cv2
import os
import openpyxl as pyxl
def getinfo(img, ws, col, index):
    greyimg = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    ws[chr(2*col+65)+str(index)] = greyimg.mean()
    ws[chr(2*col+66)+str(index)] = np.std(greyimg)
    
def readImg(filepath):
    wb = pyxl.load_workbook(filepath+"/info.xlsx")
    ws = wb.active
    imgs = []
    filenames = os.listdir(filepath)
    counter = 1
    ws['A1'] = "原亮度(0~255)"
    ws['B1'] = "原对比度(灰度图标准差)"
    for filename in filenames:
        file = os.path.join(filepath,filename)
        print(file)
        if ".xlsx" not in filename and os.path.isfile(file):
            counter += 1
            img = cv2.imdecode(np.fromfile(file,dtype=np.uint8), -1)
            getinfo(img, ws, 0, counter)
            imgs.append(img)
    wb.save(filepath+"/info.xlsx")
    return imgs

def convolution(filepath, kernels, imgs):
    wb = pyxl.load_workbook(filepath+"/info.xlsx")
    ws = wb.worksheets[0]
    for i in range(len(kernels)):
        ws[chr(2*i+67)+'1'] = "Kernel"+str(i+1)+"亮度"
        ws[chr(2*i+68)+'1'] = "Kernel"+str(i+1)+"对比度"
        for j in range(len(imgs)):
            dst = cv2.filter2D(imgs[j], -1, kernels[i])
            getinfo(dst, ws, i+1, j+2)
            if(not os.path.exists(filepath+"/kernel"+str(i+1))):
                os.mkdir(filepath+"/kernel"+str(i+1))
            cv2.imencode('.jpg', dst)[1].tofile(filepath+"/kernel"+str(i+1)+"/img"+str(j+1)+".jpg")
            #cv2.imwrite(filepath+"/kernel"+str(i+1)+"/img"+str(j+1)+".jpg", dst)
    wb.save(filepath+"/info.xlsx")

def getKernels():
    kernels = []
    ker1 = np.array([1,1,1,1,9,1,1,1,1], dtype=float)/9.0
    ker1.reshape([3,3])
    kernels.append(ker1)
    ker2 = cv2.getGaussianKernel(3, -1)
    kernels.append(ker2)
    ker3 = np.array([1,1,1,1,1,1,1,1,1], dtype=float)/9.0
    ker3.reshape([3,3])
    kernels.append(ker3)
    ker4 = np.array([-1,-1,-1,-1,9,-1,-1,-1,-1], dtype=float)
    ker4.reshape([3,3])
    kernels.append(ker4)
    ker5 = np.array([0,-1,0,-1,5,-1,0,-1,0], dtype=float)
    ker5.reshape([3,3])
    kernels.append(ker5)

    return kernels
def main():
    fileroot = input("Input filepath: ")
    kernels = getKernels()
    for dir in os.listdir(fileroot):
        filename = os.path.join(fileroot, dir)
        if(os.path.isdir(filename)):
            wb = pyxl.Workbook()
            wb.save(filename+"/info.xlsx")
            imgs = readImg(filename)
            convolution(filename, kernels, imgs)
if __name__ == "__main__":
    main()
